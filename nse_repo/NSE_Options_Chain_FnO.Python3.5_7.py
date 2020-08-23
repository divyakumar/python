
import requests	
import pandas as pd
import numpy as np
from bs4 import BeautifulSoup
import sys
import time
import re
import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from os import path
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
#from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
#color scale in python is -> #RRGGBB

'''
Base_url =("https://www.nseindia.com/live_market/dynaContent/"+
           "live_watch/option_chain/optionKeys.jsp?symbolCode=2772&symbol=UBL&"+
           "symbol=UBL&instrument=OPTSTK&date=-&segmentLink=17&segmentLink=17")
'''
#https://nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument=OPTIDX&symbol=BANKNIFTY&date=19APR2018
#https://nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?symbolCode=180&symbol=INFY&symbol=INFY&instrument=-&date=-&segmentLink=17&symbolCount=2&segmentLink=17
#https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument=OPTSTK&symbol=RELIANCE&date=31MAY2018
#https://www.nseindia.com/live_market/dynaContent/live_watch/get_quote/GetQuote.jsp?symbol=RELIANCE&illiquid=0&smeFlag=0&itpFlag=0


'''
black	#000000
blue	#0000FF
brown	#800000
cyan	#00FFFF
gray	#808080
green	#008000
lime	#00FF00
magenta	#FF00FF
navy	#000080
orange	#FF6600
pink	#FF00FF
purple	#800080
red		#FF0000
silver	#C0C0C0
white	#FFFFFF
yellow	#FFFF00
'''

#----------- functions --------------------------------------------------------
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startRow, startCol, endRow, endCol, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected
         
def pasteRange(startRow, startCol, endRow, endCol, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def pasteRange2(startRow, startCol, endRow, endCol, sheetReceiving, copiedData, replace_str, to_sheet_name, ncalendar):
	countRow = 0
	#print("paste2 replace_str=%s (r,c)=%d,%d -> (%d,%d) sheetRx=%s" %(replace_str[0], startRow, startCol, endRow, endCol, sheetReceiving));
	for i in range(startRow,endRow+1,1):
		countCol = 0
		for j in range(startCol,endCol+1,1):
			if((ncalendar>1) and (re.search('INDEX\(\[1\]', str(copiedData[countRow][countCol])))):
				sheetReceiving.cell(row = i, column = j).value = str(copiedData[countRow][countCol]).replace('[1]',replace_str[1]).replace('NIFTY.11MarEoD', to_sheet_name);
				#print("1re.search=%s replace=%s" %(re.search('INDEX\(\[1\]', str(copiedData[countRow][countCol])), str(copiedData[countRow][countCol]).replace('[1]',replace_str[1])))
			elif((ncalendar>2) and (re.search('INDEX\(\[2\]', str(copiedData[countRow][countCol])))):
				sheetReceiving.cell(row = i, column = j).value = str(copiedData[countRow][countCol]).replace('[2]',replace_str[2]).replace('NIFTY.11MarEoD', to_sheet_name);
				#print("2re.search=%s replace=%s" %(re.search('INDEX\(\[2\]', str(copiedData[countRow][countCol])), sheetReceiving.cell(row = i, column = j).value))
			elif((ncalendar>3) and (re.search('INDEX\(\[3\]', str(copiedData[countRow][countCol])))):
				sheetReceiving.cell(row = i, column = j).value = str(copiedData[countRow][countCol]).replace('[3]',replace_str[3]).replace('NIFTY.11MarEoD', to_sheet_name);
				#print("3re.search=%s replace=%s" %(re.search('INDEX\(\[3\]', str(copiedData[countRow][countCol])), str(copiedData[countRow][countCol]).replace('[3]',replace_str[3])))
			else:
				sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
			countCol += 1
		countRow += 1
		
def get_coordinates_from_cell(cell):
	xy = coordinate_from_string(cell); # returns ('A', 4)
	col = column_index_from_string(xy[0]);
	row = xy[1]
	return (row, col)
		
def get_opt_table(t_symbol, t_sheet_name, myexpiry, writer, out_file_name):
	#global writer;
	global my_optIDX_optSTK;
	global my_write_sheet_name;

	if (t_symbol=="BANKNIFTY") or (t_symbol=="NIFTY"):
		#doesnt work yet
		Base_url =("https://www.nseindia.com/live_market/dynaContent/"+
			   "live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument="+my_optIDX_optSTK+"&symbol="+str(t_symbol)+"&date="+str(myexpiry));
	elif (t_symbol=="USDINR"):
		Base_url =("https://www.nseindia.com/live_market/dynaContent/live_watch/fxTracker/optChainDataByExpDates.jsp?symbol="+str(t_symbol)+"&instrument=OPTCUR&expiryDt="+str(myexpiry))
	else:
		Base_url =("https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?segmentLink=17&instrument="+
					my_optIDX_optSTK+"&symbol="+t_symbol+"&date="+myexpiry);

	#Base_url="https://beta.nseindia.com/get-quotes/derivatives?symbol=BANKNIFTY&identifier=OPTIDXBANKNIFTY26-12-2019"
	print("url: %s" %(Base_url));
	if 1:
		#hack to fake field: User-Agent, DEC 2019
		headers = {'User-Agent': 'Chrome/79.0.3945.88'}
		page = requests.get(Base_url, headers=headers)
	else:
		page = requests.get(Base_url)
	page.status_code
	page.content

	soup = BeautifulSoup(page.content, 'html.parser')
	#print(soup.prettify())

	table_it = soup.find_all(class_="opttbldata")
	table_cls_1 = soup.find_all(id="octable")

	col_list = []

	# The code given below will pull the headers of the Option Chain table
	for mytable in table_cls_1:
		table_head = mytable.find('thead')
		
		try:
			rows = table_head.find_all('tr')
			for tr in rows: 
				cols = tr.find_all('th')
				for th in cols:
					er = th.text
					ee = er.encode('utf8')   
					ee = str(ee, 'utf-8')
					col_list.append(ee)
					
		except:
			print ("no thead")
		

	col_list_fnl = [e for e in col_list if e not in ('CALLS','PUTS','Chart','Chart ','\xc2\xa0','\xa0')];
	col_list_fnl += ['EQ-LTP'];
					
	#print (col_list_fnl)

	table_cls_2 = soup.find(id="octable")
	req_row = table_cls_2.find_all('tr')

	new_table = pd.DataFrame(index=range(0,len(req_row)-3) , columns=col_list_fnl)

	row_marker = 0 
	for row_number, tr_nos in enumerate(req_row):
		 
		# This ensures that we use only the rows with values    
		if row_number <=1 or (row_number == len(req_row)-1):
			continue

		if (t_symbol=="USDINR"):
			td_columns = tr_nos.find_all('td');
			select_cols = td_columns[1:20]
			
			for nu, column in enumerate(select_cols):
			 
				utf_string = column.get_text()
				utf_string = utf_string.strip('\n\r\t": ')
				 
				tr = utf_string.encode('utf-8')
				tr = str(tr, 'utf-8')
				tr = tr.replace(',' , '')

				try:
					new_table.iloc[row_marker,[nu]]= float(tr) if (re.fullmatch('-',tr) is None) else None;
				except ValueError as e:
					print("row_number:%s tr:%s nu:%s row_marker:%s e:%s" %(row_number, tr, nu, row_marker, e));
			row_marker += 1
		else:
			td_columns = tr_nos.find_all('td');
			# This removes the graphs columns
			select_cols = td_columns[1:22]
		
			for nu, column in enumerate(select_cols):
			 
				utf_string = column.get_text()
				utf_string = utf_string.strip('\n\r\t": ')
				 
				tr = utf_string.encode('utf-8')
				tr = str(tr, 'utf-8')
				tr = tr.replace(',' , '')

				new_table.iloc[row_marker,[nu]]= float(tr) if (re.fullmatch('-',tr) is None) else None;
			row_marker += 1

	#----- >>> my additions -------------------------------------------------------
	t_data = pd.DataFrame([], columns=col_list_fnl);
	t_columns=col_list_fnl;

	t_data.to_excel(writer, sheet_name=t_sheet_name, columns=t_columns, index=False,header=True); #needs to be inited for conditional format writing
	worksheet = writer.sheets[t_sheet_name];

	if my_optIDX_optSTK=="OPTCUR":
		table_eq_info = soup.find_all(width="67%")[0];
	else:
		table_eq_info = soup.find_all(width="100%")[0];
	my_eq_info_list = ([])
	t_eq_ltp=0.0;
	try:
		if my_optIDX_optSTK=="OPTCUR":
			t_eq_ltp = float(table_eq_info.get_text().split("IST :")[-1].strip(' '));
			t_str1 = str(table_eq_info);
			t_date_time = t_str1[t_str1.find("as on")+6:t_str1.find("IST")-1];
			t_time = t_date_time[-8:];
		else:
			for th in table_eq_info:
				er = th.text
				ee = er.encode('utf8')   
				ee = str(ee, 'utf-8')
				my_eq_info_list.append(ee)
			t_str1 = str(my_eq_info_list);
			if my_optIDX_optSTK=="OPTIDX":
				t_name_ltp = t_str1[t_str1.find("Index: ")+7:t_str1.find('\\xa0')];
			else:
				t_name_ltp = t_str1[t_str1.find("Stock: ")+7:t_str1.find('\\xa0')];
			t_eq_ltp = float(t_name_ltp.split()[1]);
		
			t_date_time = t_str1[t_str1.find("As on")+6:t_str1.find("IST")-1];
			t_time = t_date_time[-8:];
		print("LTP=%s Updated TIME=%s" %(t_eq_ltp, t_time));
	except:
		print ("no eq_info");
		
	new_table.iloc[0, -1] = t_eq_ltp; #ADD LTP to the last column, 2nd row

	#===POST-PROCESS write to excel===========================================================
	# Convert the dataframe to an XlsxWriter Excel object.
	t_columns = col_list_fnl;

	#--- color strike steps---
	from_col = np.where(pd.DataFrame(t_columns) == 'Strike Price')[0][0];
	#print ("strikes=%s" %(new_table['Strike Price']));
	to_col = from_col;
	from_row = 1;
	to_row = len(new_table);

	coord_from = worksheet.cell(row=1+from_row,column=1+from_col).coordinate;
	coord_to   = worksheet.cell(row=1+to_row,column=1+to_col).coordinate;
	worksheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=greyFill))
	
	#--- color ITM calls ---
	from_col = 0;
	to_col = np.where(pd.DataFrame(t_columns) == 'Strike Price')[0][0];
	from_row = 1;
	to_row = [k for k in range(0, len(new_table)) if (float(new_table['Strike Price'][k]) > float(t_eq_ltp) ) ][0];
	
	coord_from = worksheet.cell(row=1+from_row,column=1+from_col).coordinate;
	coord_to   = worksheet.cell(row=1+to_row,column=1+to_col).coordinate;
	worksheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=yellowFill))
	
	#--- color ITM puts ---
	from_col = np.where(pd.DataFrame(t_columns) == 'Strike Price')[0][0];
	to_col = len(t_columns)-1;
	from_row = [k for k in range(0, len(new_table)) if (float(new_table['Strike Price'][k]) > float(t_eq_ltp) ) ][0] + 1;
	to_row = len(new_table);
	
	coord_from = worksheet.cell(row=1+from_row,column=1+from_col).coordinate;
	coord_to   = worksheet.cell(row=1+to_row,column=1+to_col).coordinate;
	worksheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=yellowFill))
	
	new_table.to_excel(writer, sheet_name=t_sheet_name, columns=t_columns, index=False, header=True);
	print("written to file:%s, sheet:%s" %(out_file_name, t_sheet_name));
	#----- <<< my additions -------------------------------------------------------

def get_time():
	t_z = time.ctime().split();
	t_date = t_z[2]+t_z[1];
	t_hour = time.ctime().split()[3].replace(":","_");
	[t_hr, t_min, t_sec] = time.ctime().split()[3].split(':')
	return [t_hr, t_min, t_sec, t_date, t_z];

def get_write_sheet_name(mysymbol):
	[t_hr, t_min, t_sec, t_date, t_z] = get_time();
	if (((int(t_hr) > 17) and (int(t_min) >= 30)) or (int(t_hr) > 18)):
		t_sheet_name = mysymbol+"."+t_z[2]+t_z[1]+"EoD";
	#elif (((int(t_hr) < 10) and (int(t_min) < 15)) or (int(t_hr) < 9)):
	elif (int(t_hr) < 9):
		t_sheet_name = mysymbol+"."+t_z[2]+t_z[1]+"EoD";
	else:
		t_sheet_name = mysymbol+"."+t_z[2]+t_z[1]+t_hr+"_"+t_min;
	return t_sheet_name;

def get_excel_writer_out_file_name(mysymbol, myexpiry, myoutfilename=''):
	[t_hr, t_min, t_sec, t_date, t_z] = get_time();
	if (myoutfilename is not ''):
		out_file_name = myoutfilename;
		if(path.isfile(myoutfilename)):
			out_file_name = myoutfilename;
			print("Appending sheet: %s to file %s" %(t_sheet_name, out_file_name))
			
			book = load_workbook(out_file_name)
			writer = pd.ExcelWriter(out_file_name, engine = 'openpyxl')
			writer.book = book
		else:
			writer = pd.ExcelWriter(out_file_name, engine='openpyxl'); #add sheet to new file
	elif (((t_hr > '15') and (t_min > '29')) or (t_hr > '16')):
		out_file_name = "Option_Chain_FnO."+mysymbol+"_"+myexpiry+"."+t_date+"EoD"+".xlsx";
		writer = pd.ExcelWriter(out_file_name, engine='openpyxl'); #add sheet to new file
	elif ((t_hr < '9')):
		yesterday = datetime.date.today() - datetime.timedelta(days = 1)
		t_date = yesterday.ctime().split(' ')[2];
		out_file_name = "Option_Chain_FnO."+mysymbol+"_"+myexpiry+"."+t_date+"EoD"+".xlsx";
		writer = pd.ExcelWriter(out_file_name, engine='openpyxl'); #add sheet to new file
	else:
		out_file_name = "Option_Chain_FnO."+mysymbol+"_"+myexpiry+"."+t_date+t_hr+"_"+t_min+".xlsx";
		writer = pd.ExcelWriter(out_file_name, engine='openpyxl'); #add sheet to new file
	return (writer, out_file_name);
	
def copy_formulas_step1(to_sheet, sym, i):
	#File to be copied from

	if (sym=="USDINR"):
		wb = load_workbook("usdinr_template.2strike_delta.2.xlsx")
		from_sheet = wb["USDINR.NearWeek"]
		to_row = 45;
	elif i:
		wb = load_workbook("nifty_template.2strike1strike_delta.v5.xlsx")
		from_sheet = wb["NIFTY.month2strike"]
		to_row = 120;
	else:
		wb = load_workbook("nifty_template.2strike1strike_delta.v5.xlsx")
		from_sheet = wb["NIFTY.week1strike"]
		to_row = 120;
	
	print("Copying formulas...")
	(from_row, from_col) = get_coordinates_from_cell('w1');
	(_, to_col) = get_coordinates_from_cell('ar1');
	
	#to_col = from_col + 4;
	selectedRange = copyRange(from_row, from_col, to_row, to_col, from_sheet) #Change the 4 number values
	pastingRange = pasteRange(from_row, from_col, to_row, to_col, to_sheet, selectedRange) #Change the 4 number values
	
	coord_from = 'w1'; 
	coord_to   = 'z120';
	to_sheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=lightgreyFill))
	
	coord_from = 'aa1'; 
	coord_to   = 'ad120';
	to_sheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=cloudyblueFill))

	#copy openInterest sum v3:Call v4:put v5:PCR (putOI/CallOI)
	(from_row, from_col) = get_coordinates_from_cell('v3');
	(_, to_col) = get_coordinates_from_cell('v5');
	selectedRange = copyRange(from_row, from_col, to_row, to_col, from_sheet) #Change the 4 number values
	pastingRange = pasteRange(from_row, from_col, to_row, to_col, to_sheet, selectedRange) #Change the 4 number values

	#shade openInterest sum v3:Call v4:put v5:PCR (putOI/CallOI)
	coord_from = 'v3'; 
	coord_to   = 'v5';
	to_sheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=lightgreyFill))
    
	print("Range copied and pasted!")
	wb.close();

def copy_formulas_step2(to_sheet, to_sheet_name, out_file_name, ncalendar):
	wbr = load_workbook("nifty_template.2strike1strike_delta.v5.xlsx")
	from_sheet = wbr["NIFTY.week1strike"]
	replace_str = [];
	for i in range(len(out_file_name)-1):
		replace_str += ["["+out_file_name[i]+ "]"]

	print("2.Copying formulas...")
	(from_row, from_col) = get_coordinates_from_cell('af13');
	(to_row, to_col) = get_coordinates_from_cell('ar120');
	selectedRange = copyRange(from_row, from_col, to_row, to_col, from_sheet)

	#print("from_row=%s, from_col=%s, to_row=%s, to_col=%s, to_sheet=%s, replace_str=%s, to_sheet_name=%s" %(from_row, from_col, to_row, to_col, to_sheet, replace_str, to_sheet_name));
	pastingRange = pasteRange2(from_row, from_col, to_row, to_col, to_sheet, selectedRange, replace_str, to_sheet_name, ncalendar)
	
	#add formatting
	coord_from = 'af13'; 
	coord_to   = 'ar120';
	to_sheet.conditional_formatting.add(coord_from+":"+coord_to, CellIsRule(operator='between', formula=[1e9,-1e9], stopIfTrue=True, fill=cloudyblueFill))
	
	print("2.Range copied and pasted!")
	wbr.close();

#----------------------- Initializations --------------------------------------
yellowFill = PatternFill(start_color='fffcc4', end_color='fffcc4', fill_type='solid')
greyFill = PatternFill(start_color='8F8F8F', end_color='8F8F8F', fill_type='solid')
skyblueFill = PatternFill(start_color='75bbfd', end_color='75bbfd', fill_type='solid')
cloudyblueFill = PatternFill(start_color='acc2d9', end_color='acc2d9', fill_type='solid')
cementFill = PatternFill(start_color='a5a391', end_color='a5a391', fill_type='solid')
lightgreyFill = PatternFill(start_color='d8dcd6', end_color='d8dcd6', fill_type='solid')

mysymbol='';
myexpiry='29AUG2019';
my_optIDX_optSTK="OPTSTK";
myoutfilename=''
myformat=list([]);

myargs = sys.argv[1:];
print("Usage: <>.py <mysymbol> <myexpiry/%s> <outfilename>" %(myexpiry));
print("myargs=%s" %(myargs));

if (mysymbol=="BANKNIFTY") or (mysymbol=="NIFTY"):
	my_optIDX_optSTK="OPTIDX";
elif (mysymbol=="USDINR"):
	my_optIDX_optSTK="OPTCUR";

#==============================================================================
if myargs:
	#tested part
	myargs1 = myargs;
	sym = myargs1[0];
	if (sym=="BANKNIFTY") or (sym=="NIFTY"):
		my_optIDX_optSTK="OPTIDX";
	elif (sym=="USDINR"):
		my_optIDX_optSTK="OPTCUR";
	else:
		my_optIDX_optSTK="OPTSTK";

	t_sheet_name = get_write_sheet_name(sym);
	writer = list(range(len(myargs1)));
	out_file_name = list(range(len(myargs1)));
	for i in range(len(myargs1)-1): #usage is "<sym> <exp-near> <exp-mid> <exp-far>
		myexpiry = myargs1[i+1].upper();
		myoutfilename = "Option_Chain_FnO."+sym+"_"+myexpiry+".xlsx";
		writer[i], out_file_name[i] = get_excel_writer_out_file_name(sym, myexpiry, myoutfilename)
		print("sym:%s nExpiry:%s out:%s sheet:%s" %(sym, myexpiry, out_file_name[i], t_sheet_name))
		get_opt_table(sym, t_sheet_name, myexpiry, writer[i], out_file_name[i]);
		copy_formulas_step1(writer[i].sheets[t_sheet_name], sym, i)
		
		writer[i].book.active.sheet_view.zoomScale = 70
		writer[i].book.active.freeze_panes = writer[i].book.active['a4']
		writer[i].save();
		writer[i].close();
		
	book = load_workbook(out_file_name[0]);
	writer[0] = pd.ExcelWriter(out_file_name[0], engine = 'openpyxl')
	writer[0].book = book;
	#print("step2 %s %s out:%s %s worksheet=%s" %(sym, myexpiry, out_file_name[0], t_sheet_name, writer[0].book.worksheets[-1]))
	copy_formulas_step2(writer[0].book.worksheets[-1], writer[0].book.sheetnames[-1], out_file_name, len(myargs1[1:]))
	
	writer[0].book.active.sheet_view.zoomScale = 70
	writer[0].book.active.freeze_panes = writer[0].book.active['a4']	
	writer[0].save();
	writer[0].close();
else:
	#untested part
	with open('FNO_list.txt','r') as fno_list:
		flines = fno_list.readlines();
		for line1 in flines:
			print("%s " %(line1),end='')
			line1 = line1.split('\n')[0];
			myargs1 = line1.split(',');
			sym = myargs1[0];
			if (sym=="BANKNIFTY") or (sym=="NIFTY"):
				my_optIDX_optSTK="OPTIDX";
			elif (sym=="USDINR"):
				my_optIDX_optSTK="OPTCUR";
			else:
				my_optIDX_optSTK="OPTSTK";
			t_sheet_name = get_write_sheet_name(sym);
			writer = list(range(len(myargs1)));
			out_file_name = list(range(len(myargs1)));
			if(len(myargs1) > 1):
				for i in range(len(myargs1)-1):
					myexpiry = myargs1[i+1].upper();
					myoutfilename = "Option_Chain_FnO."+sym+"_"+myexpiry+".xlsx";
					writer[i], out_file_name[i] = get_excel_writer_out_file_name(sym, myexpiry, myoutfilename)
					print("multiExpiry in:%s %s out:%s %s" %(sym, myexpiry, out_file_name[i], t_sheet_name))
					get_opt_table(sym, t_sheet_name, myexpiry, writer[i], out_file_name[i]);
					copy_formulas_step1(writer[i].sheets[t_sheet_name], sym, i)

					writer[i].book.active.sheet_view.zoomScale = 80
					writer[i].book.active.freeze_panes = writer[i].book.active['a4']
					writer[i].save();
					writer[i].close();
			else:
				myoutfilename = "Option_Chain_FnO."+sym+"_"+myexpiry+".xlsx";
				writer[i], out_file_name[i] = get_excel_writer_out_file_name(sym, myexpiry, myoutfilename)
				print("singleExpiry in:%s %s out:%s %s" %(sym, myexpiry, out_file_name[i], t_sheet_name))
				get_opt_table(sym, t_sheet_name, myexpiry, writer[i], out_file_name[i]);
				copy_formulas_step1(writer[i].sheets[t_sheet_name], sym, i)
				writer[i].book.active.sheet_view.zoomScale = 80
				writer[i].book.active.freeze_panes = writer[i].book.active['a4']
				writer[i].save();
				writer[i].close();
#==============================================================================
